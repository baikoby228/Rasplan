from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from datetime import *
import threading
import openpyxl 
import copy
import time
import random
import math
import sys
import os

window = Tk()

window.geometry('312x338')
window.resizable(False, False)
window.title("Rasplan")

if getattr(sys, 'frozen', False):
    icon_path = os.path.join(sys._MEIPASS, 'icon.ico')
else:
    icon_path = os.path.join(os.path.dirname(__file__), 'icon.ico')
window.iconbitmap(icon_path)

def clear_window():
    for widget in window.winfo_children():
        widget.destroy()
    time.sleep(1)

def is_full_None(a):
    for x in a:
        if x != None:
            return False
    return True

flag = [False, False, False, False, False]

data_teachers = []
def open_excel1():
    file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
    if file_path:
        wb = openpyxl.load_workbook(file_path)  
        sheet = wb.active

        flag[0] = True
        data_teachers.clear()
        for row in sheet.iter_rows(values_only=True):
            a = list(row)
            if not is_full_None(a):
                data_teachers.append(a)

data_classes = []
def open_excel2():
    file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
    if file_path:
        wb = openpyxl.load_workbook(file_path)  
        sheet = wb.active

        flag[1] = True
        data_classes.clear()
        for row in sheet.iter_rows(values_only=True):
            a = list(row)
            if not is_full_None(a):
                data_classes.append(a)

data_start = []
def open_excel3():
    file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
    if file_path:
        wb = openpyxl.load_workbook(file_path)  
        sheet = wb.active

        flag[2] = True
        data_start.clear()
        for row in sheet.iter_rows(values_only=True):
            a = list(row)
            if not is_full_None(a):
                data_start.append(a)

data_merge_lessons = []
def open_excel4():
    file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
    if file_path:
        wb = openpyxl.load_workbook(file_path)  
        sheet = wb.active

        flag[3] = True
        data_merge_lessons.clear()
        for row in sheet.iter_rows(values_only=True):
            a = list(row)
            if not is_full_None(a):
                data_merge_lessons.append(a)

data_enemy_lessons = []
def open_excel5():
    file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])
    if file_path:
        wb = openpyxl.load_workbook(file_path)  
        sheet = wb.active

        flag[4] = True
        data_enemy_lessons.clear()
        for row in sheet.iter_rows(values_only=True):
            a = list(row)
            if not is_full_None(a):
                data_enemy_lessons.append(a)

def encrypt(number, letter):
    return str(number) + '|' + letter

def restore(s):
    flag = False
    number = 0
    letter = ''
    for i in range(len(s)):
        if s[i] == '|':
            flag = True
            continue

        if not flag:
            number = number * 10 + int(s[i])
        else:
            letter = s[i]
            return number, letter

classes_in_parallel = {}
def build_classes_in_parallel():
    c = 0
    while c < len(data_classes[0]):
        classes_in_parallel[data_classes[0][c]] = []
        for i in range(0, len(data_classes[0][c + 1])):
            if 'A' <= data_classes[0][c + 1][i] and data_classes[0][c + 1][i] <= 'Я':
                classes_in_parallel[data_classes[0][c]].append(data_classes[0][c + 1][i])
        c += 2

    for number, value in classes_in_parallel.items():
        for letter in value:
            teachers[encrypt(number, letter)] = {}

teachers = {}
def build_teachers():
    r = 0
    while r < len(data_teachers):
        teacher_name = data_teachers[r][0]
        while r < len(data_teachers) and (data_teachers[r][0] == teacher_name or data_teachers[r][0] == None):
            subject_name = data_teachers[r][1]
            while r < len(data_teachers) and (data_teachers[r][0] == teacher_name or data_teachers[r][0] == None) and (data_teachers[r][1] == subject_name or data_teachers[r][1] == None):
                number = data_teachers[r][2]
                for i in range(len(data_teachers[r][3])):
                    if 'А' <= data_teachers[r][3][i] and data_teachers[r][3][i] <= 'Я':
                        if subject_name not in teachers[encrypt(number, data_teachers[r][3][i])]:
                            teachers[encrypt(number, data_teachers[r][3][i])][subject_name] = []
                        teachers[encrypt(number, data_teachers[r][3][i])][subject_name].append(teacher_name)
                r += 1

lessons_per_week = {}
def build_lessons_per_week():   
    c = 0
    while c < len(data_classes[0]):
        number = data_classes[0][c]
        lessons_per_week[number] = {}

        r = 1
        while r < len(data_classes) and data_classes[r][c] != None:
            lessons_per_week[number][data_classes[r][c]] = data_classes[r][c + 1]
            r += 1
        c += 2

names_of_teachers = []
def build_names_of_teachers():
    for i in range(len(data_teachers)):
        if data_teachers[i][0] != None:
            names_of_teachers.append(data_teachers[i][0])

start_lessons = {}
def build_start_lessons():
    for c in range(len(data_start[0])):
        start_lessons[data_start[0][c]] = data_start[1][c]

merge_lessons = []
def build_merge_lessons():
    for i in data_merge_lessons[0]:
        if i != None:
            merge_lessons.append(i)

enemy_lessons = []
def build_enemy_lessons():
    for i in range(len(data_enemy_lessons)):
        if data_enemy_lessons[i][0] != None and data_enemy_lessons[i][1] != None:
            enemy_lessons.append([data_enemy_lessons[i][0], data_enemy_lessons[i][1]])
            enemy_lessons.append([data_enemy_lessons[i][1], data_enemy_lessons[i][0]])

class WeeklySchedule:
    def __init__(self, number, letter):
        self.number_lessons_per_week = 0
        for key, value in lessons_per_week[number].items():
            self.number_lessons_per_week += value
        self.lessons_per_day = self.number_lessons_per_week // 5

        self.lessons = [[], [], [], [], []]
        cur = []
        for key, value in lessons_per_week[number].items():
            for i in range(value):
                cur.append(key)

        random.shuffle(cur)

        for i in range(5):
            for j in range(self.lessons_per_day):
                self.lessons[i].append(cur[self.lessons_per_day * i + j])

        if self.number_lessons_per_week % 5 != 0:
            for i in range(5):
                if self.lessons_per_day * 5 + i < self.number_lessons_per_week:
                    self.lessons[i].append(cur[self.lessons_per_day * 5 + i])
                else:
                    self.lessons[i].append(None)
            self.lessons_per_day += 1

last = []
def calc():
    last.clear()
    global schedule
    cnt = {}
    for teacher_name in names_of_teachers:
        cnt[teacher_name] = []
        for i in range(5):
            cnt[teacher_name].append([0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])

    for key, ws in schedule.items():
        number, letter = restore(key)
        for i in range(5):
            for j in range(ws.lessons_per_day):
                if ws.lessons[i][j] == None:
                    continue
                for teacher_name in teachers[key][ws.lessons[i][j]]:
                    cnt[teacher_name][i][j + start_lessons[number] - 1] += 1

    res = 0
    for teacher_name in names_of_teachers:
        for i in range(5):
            for x in cnt[teacher_name][i]:
                if x > 1:
                    res += x * 10000

    for key, ws in schedule.items():
        number, letter = restore(key)
        for day in range(5):
            cur = 0
            for i in range(ws.lessons_per_day):
                if ws.lessons[day][i] in merge_lessons:
                    if i != 0 and ws.lessons[day][i - 1] == ws.lessons[day][i] or i != ws.lessons_per_day - 1 and ws.lessons[day][i + 1] == ws.lessons[day][i]:
                        if i != 0 and ws.lessons[day][i - 1] == ws.lessons[day][i] and i != ws.lessons_per_day - 1 and ws.lessons[day][i + 1] == ws.lessons[day][i]:
                            res += 500
                    else:
                        res += 50

                for j in range(i + 1, ws.lessons_per_day):
                    if ws.lessons[day][i] == ws.lessons[day][j]:
                        last.append(ws.lessons[day][i])
                        cur += 1
                    if [ws.lessons[day][i], ws.lessons[day][j]] in enemy_lessons:
                        res += 2500
            res += (2 * cur) ** 2

    for i in names_of_classes:  
        number, letter = restore(i)
        for key, value in lessons_per_week[number].items():
            if value > 5:
                res -= value - 5
    return res

rnd = ''
data = []
def out():
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    for row in data:
        sheet.append(row)

    for c in sheet[1]:
        c.font = openpyxl.styles.Font(bold=True)
        c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    for col in sheet.columns:
        sz = 0
        col = list(col)
        for c in col:
            try:
                if len(str(c.value)) > sz:
                    sz = len(c.value)
            except:
                pass
        sheet.column_dimensions[col[0].column_letter].width = sz + 2

    now = datetime.now()
    global rnd
    rnd = ''
    for i in range(4):
        rnd += str(random.randint(0, 9))
    if getattr(sys, 'frozen', False):
        script_dir = os.path.dirname(sys.executable)
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
    save_path = os.path.join(script_dir, f"schedule_{now.date()}_{rnd}.xlsx")
    wb.save(save_path)

finish = False
schedule = {}
start_schedule = {}
names_of_classes = []
def generation():
    ans = calc()
    best = ans
    
    global schedule
    best_schedule = copy.deepcopy(schedule)

    t = 1
    start_time = time.time()
    while ans != 0 and time.time() - start_time < 30:
        t *= 0.99

        old_schedule = copy.deepcopy(schedule)
        class_name = names_of_classes[random.randint(0, len(names_of_classes) - 1)]

        day1 = random.randint(0, 4)
        num1 = random.randint(0, schedule[class_name].lessons_per_day - 1)
        day2 = random.randint(0, 4)
        num2 = random.randint(0, schedule[class_name].lessons_per_day - 1)

        if (schedule[class_name].lessons[day1][num1] != None and schedule[class_name].lessons[day2][num2] != None) or num1 == num2:
            schedule[class_name].lessons[day1][num1], schedule[class_name].lessons[day2][num2] = schedule[class_name].lessons[day2][num2], schedule[class_name].lessons[day1][num1]
            cur = calc()
            if cur < ans or random.random() < math.exp(-abs(cur - ans) / t):
                ans = cur
            else:
                schedule = old_schedule
            if ans < best:
                best = ans
                best_schedule = schedule
    global is_generating
    for i in last:
        if i not in merge_lessons:
            is_generating = False
            return

    global finish
    finish = True

    sz = 0
    for key, ws in best_schedule.items():
        sz = max(sz, ws.lessons_per_day)
    
    for class_name in range(1 + 5 * (sz + 1)):
        data.append([])
    
    for i in range(len(names_of_classes)):
        class_name = names_of_classes[i]
        number, letter = restore(class_name)
        data[0].append(str(number) + letter)

        for j in range(5):
            for u in range(best_schedule[class_name].lessons_per_day):
                data[1 + j * (sz + 1) + u].append(best_schedule[class_name].lessons[j][u])
            for u in range(best_schedule[class_name].lessons_per_day, sz + 1):
                data[1 + j * (sz + 1) + u].append(None)
        if i != len(names_of_classes) - 1:
            for j in range(len(data)):
                data[j].append(None)
    out()
    is_generating = False

def pre_main1():
    if not flag[0] or not flag[1]:
        return
    clear_window()
    window.update_idletasks()
    window.after(1, pre_main2)

def pre_main2():
    style = ttk.Style()
    style.configure('1.TLabel', font=('Helvetica', 37, 'bold'), relief='flat')
    label_info = ttk.Label(window, style='1.TLabel', text='      Идёт\nгенерация...')
    label_info.grid(padx=10, pady=20)
    window.update_idletasks()
    window.after(1, main(label_info))

def check_generation_status(label_info):
    if is_generating:
        window.after(500, check_generation_status, label_info)
    else:
        global finish
        if not finish:
            global schedule, start_schedule
            schedule = copy.deepcopy(start_schedule)
            start_generation(label_info)
        else:
            label_info.config(text=' Генерания\n завершена\n      ' + rnd)

is_generating = False
def start_generation(label_info):
    global is_generating
    is_generating = True
    thread = threading.Thread(target=generation)
    thread.start()
    check_generation_status(label_info)

def main(label_info):
    build_classes_in_parallel()
    build_teachers()
    build_lessons_per_week()
    build_names_of_teachers()

    if flag[2]:
        build_start_lessons()
    else:
        for i in range(5, 12):
            start_lessons[i] = 1
    if flag[3]:
        build_merge_lessons()
    if flag[4]:
        build_enemy_lessons()

    global schedule
    for key, value in classes_in_parallel.items():
        for letter in value:
            schedule[encrypt(key, letter)] = WeeklySchedule(key, letter)
            names_of_classes.append(encrypt(key, letter))
    global start_schedule
    start_schedule = copy.deepcopy(schedule)

    global finish
    schedule = copy.deepcopy(start_schedule)
    start_generation(label_info)

style0 = ttk.Style()
style0.configure('0.TButton', font=('Helvetica', 100, 'bold'), padding=0, relief='flat')

open_button1 = ttk.Button(window, text='+', style='0.TButton', command=open_excel1)
open_button1.config(width=2)
open_button1.grid(row=0, column=0, sticky='w')

open_button2 = ttk.Button(window, text='+', style='0.TButton', command=open_excel2)
open_button2.config(width=2)
open_button2.grid(row=0, column=1, sticky='w')

style1 = ttk.Style()
style1.configure('1.TButton', font=('Helvetica', 65, 'bold'), padding=0, relief='flat')

open_button3 = ttk.Button(window, text='+', style='1.TButton', command=open_excel3)
open_button3.config(width=2)
open_button3.grid(row=1, column=0, columnspan=2, sticky='w')

open_button4 = ttk.Button(window, text='+', style='1.TButton', command=open_excel4)
open_button4.config(width=2)
open_button4.grid(row=1, column=0, columnspan=2)

open_button5 = ttk.Button(window, text='+', style='1.TButton', command=open_excel5)
open_button5.config(width=2)
open_button5.grid(row=1, column=0, columnspan=2, sticky='e')

style2 = ttk.Style()
style2.configure('2.TButton', font=('Helvetica', 30, 'bold'), padding=0, relief='flat')

generation_button = ttk.Button(window, text='Сгенерировать', style='2.TButton', command=pre_main1)
generation_button.grid(row=2, column=0, columnspan=2, ipady=7)

window.mainloop()
